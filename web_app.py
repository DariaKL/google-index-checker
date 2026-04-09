#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Google Index Checker — Flask + Playwright headless Chromium + SSE
Optimized for Render free tier (512MB RAM).
"""

import csv
import io
import json
import logging
import os
import queue as _queue_mod
import sys
import time
import uuid
import threading
from datetime import datetime
from urllib.parse import quote

from flask import Flask, render_template, request, jsonify, Response, send_file
from playwright.sync_api import sync_playwright

# ── Logging ──
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    stream=sys.stderr,
)
log = logging.getLogger("idx")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

app = Flask(__name__, template_folder="templates")

# ── In-memory task storage ──
tasks = {}
tasks_lock = threading.Lock()


class CheckTask:
    def __init__(self, urls, delay_ms, lang):
        self.id = str(uuid.uuid4())[:8]
        self.urls = urls
        self.delay_ms = delay_ms
        self.lang = lang
        self.results = []
        self.status = "pending"
        self.current = 0
        self.total = len(urls)
        self.events = []
        self.stop_event = threading.Event()
        self.waiting_for_manual = threading.Event()
        self.manual_result = None

    def push_event(self, data):
        self.events.append(data)


# ═══════════════════════════════════════════════════════
#  Playwright — single dedicated worker thread
# ═══════════════════════════════════════════════════════

_pw_queue = _queue_mod.Queue()
_pw_ready = threading.Event()
_pw_alive = False

_pw_thread = None
_pw_thread_lock = threading.Lock()

# Chromium flags optimized for low-memory containers
_LAUNCH_ARGS = [
    "--no-sandbox",
    "--disable-dev-shm-usage",
    "--disable-gpu",
    "--disable-blink-features=AutomationControlled",
    "--disable-extensions",
    "--disable-infobars",
    "--disable-background-networking",
    "--disable-default-apps",
    "--disable-sync",
    "--disable-translate",
    "--metrics-recording-only",
    "--no-first-run",
]


def _pw_worker():
    """Owns Playwright + Chromium; processes jobs via _pw_queue."""
    global _pw_alive

    log.info("pw_worker: starting Playwright…")
    try:
        pw = sync_playwright().start()
    except Exception as exc:
        log.error("pw_worker: Playwright start failed: %s", exc)
        _pw_ready.set()
        return

    def launch_browser():
        log.info("pw_worker: launching Chromium…")
        b = pw.chromium.launch(headless=True, args=_LAUNCH_ARGS)
        log.info("pw_worker: Chromium OK")
        return b

    try:
        browser = launch_browser()
    except Exception as exc:
        log.error("pw_worker: Chromium launch FAILED: %s", exc)
        _pw_alive = False
        _pw_ready.set()
        pw.stop()
        return

    _pw_alive = True
    _pw_ready.set()
    log.info("pw_worker: ready for jobs")

    while True:
        job = _pw_queue.get()
        if job is None:
            break
        url, lang, holder, done_evt = job
        log.info("pw_worker: check %s", url)
        try:
            if not browser.is_connected():
                log.warning("pw_worker: browser disconnected, relaunching")
                browser = launch_browser()
            result = _do_check(browser, url, lang)
            holder["result"] = result
            log.info("pw_worker: %s → %s", url, result[0])
            if result[0] == "captcha":
                try:
                    browser.close()
                except Exception:
                    pass
                browser = launch_browser()
        except Exception as exc:
            log.error("pw_worker: error for %s: %s", url, exc)
            holder["result"] = ("error", "", f"Помилка: {str(exc)[:120]}")
            try:
                browser.close()
            except Exception:
                pass
            try:
                browser = launch_browser()
            except Exception as exc2:
                log.error("pw_worker: relaunch failed: %s", exc2)
                _pw_alive = False
        done_evt.set()

    try:
        browser.close()
    except Exception:
        pass
    pw.stop()


def _do_check(browser, url, lang):
    """Run one  site:<url>  search in headless Chromium."""
    clean = url.strip()
    if not clean.startswith(("http://", "https://")):
        clean = "https://" + clean

    search_url = (
        f"https://www.google.com/search?q=site:{quote(clean)}&num=10&hl={quote(lang)}"
    )

    context = browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        locale=f"{lang}-UA",
        viewport={"width": 1280, "height": 720},
    )
    try:
        context.add_cookies([
            {"name": "CONSENT", "value": "YES+cb.20231008-08-p0.uk+FX+684",
             "domain": ".google.com", "path": "/"},
            {"name": "SOCS",
             "value": "CAISHAgBEhJnd3NfMjAyMzEwMDgtMF9SQzIaAmVuIAEaBgiA0JyaBg",
             "domain": ".google.com", "path": "/"},
        ])

        page = context.new_page()
        page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
        page.wait_for_timeout(3000)

        current_url = page.url

        if "/sorry/" in current_url:
            return "captcha", "", "CAPTCHA — потрібна ручна перевірка"

        text = page.inner_text("body").lower()

        if "unusual traffic" in text or "our systems have detected unusual" in text:
            return "captcha", "", "Незвичний трафік — потрібна ручна перевірка"

        no_phrases = [
            "did not match any documents",
            "your search did not match",
            "no results found",
            "не відповідає жодному",
            "не збігається жодному",
            "жодних результатів",
            "жодного документа",
            "не знайдено жодного",
            "немає результатів",
        ]
        for ph in no_phrases:
            if ph in text:
                return "not_indexed", "", "Сторінка не в індексі Google"

        stats = page.query_selector("#result-stats")
        if stats:
            h3 = page.query_selector("h3")
            title = h3.inner_text()[:120] if h3 else ""
            return "indexed", title, stats.inner_text()[:120]

        h3s = page.query_selector_all("h3")
        if h3s:
            title = h3s[0].inner_text()[:120]
            return "indexed", title, f"Знайдено {len(h3s)} результат(ів)"

        if "search console" in text:
            return "not_indexed", "", "Google рекомендує Search Console (не в індексі)"

        if len(text) > 500:
            return "not_indexed", "", "Результатів не знайдено"

        return "error", "", "Не вдалося розпізнати відповідь Google"
    finally:
        context.close()


# ── Thread lifecycle ──

def _ensure_pw_worker():
    """Lazily start the Playwright worker thread (survives gunicorn fork)."""
    global _pw_thread, _pw_alive
    if _pw_thread is not None and _pw_thread.is_alive():
        return
    with _pw_thread_lock:
        if _pw_thread is not None and _pw_thread.is_alive():
            return
        log.info("ensure_pw: spawning worker thread")
        _pw_alive = False
        _pw_ready.clear()
        _pw_thread = threading.Thread(target=_pw_worker, daemon=True)
        _pw_thread.start()


def check_indexed_playwright(url, lang="uk"):
    """Submit a check to the Playwright thread and wait for result."""
    _ensure_pw_worker()
    if not _pw_ready.wait(timeout=90):
        log.error("check: Playwright not ready after 90s")
        return ("error", "", "Playwright не запустився (таймаут)")
    if not _pw_alive:
        log.error("check: Chromium didn't launch")
        return ("error", "", "Chromium не запустився на сервері")
    if _pw_thread is None or not _pw_thread.is_alive():
        log.error("check: worker thread dead")
        return ("error", "", "Worker thread зупинився")
    holder = {}
    done = threading.Event()
    _pw_queue.put((url, lang, holder, done))
    done.wait(timeout=120)
    return holder.get("result", ("error", "", "Таймаут перевірки (120с)"))


# ═══════════════════════════════════════════
#  Background check worker
# ═══════════════════════════════════════════

def run_check(task):
    task.status = "running"
    task.push_event({"type": "started", "total": task.total})

    for i, url in enumerate(task.urls):
        if task.stop_event.is_set():
            break

        task.current = i + 1

        try:
            status, title, comment = check_indexed_playwright(url, task.lang)
        except Exception as exc:
            status, title, comment = "error", "", str(exc)[:120]

        if status == "captcha":
            task.manual_result = None
            task.waiting_for_manual.clear()
            task.push_event({
                "type": "need_manual",
                "url": url,
                "num": i + 1,
                "total": task.total,
                "comment": comment,
            })
            task.waiting_for_manual.wait(timeout=300)

            if task.stop_event.is_set():
                break

            if task.manual_result:
                status = task.manual_result.get("status", "error")
                title = task.manual_result.get("title", "")
                comment = task.manual_result.get("comment", "Перевірено вручну")
                task.manual_result = None
            else:
                status = "error"
                comment = "Таймаут ручної перевірки"

        row = {
            "num": i + 1, "url": url, "status": status,
            "title": title, "comment": comment,
        }
        task.results.append(row)
        task.push_event({
            "type": "result", "row": row,
            "current": i + 1, "total": task.total,
        })

        if i < task.total - 1 and not task.stop_event.is_set():
            time.sleep(task.delay_ms / 1000)

    task.status = "done"
    task.push_event({"type": "done"})


# ═══════════════════════════════════════════
#  Flask routes
# ═══════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/start", methods=["POST"])
def start_check():
    data = request.get_json()
    urls_text = data.get("urls", "")
    delay_ms = int(data.get("delay", 3000))
    lang = data.get("lang", "uk")

    lines = [ln.strip() for ln in urls_text.splitlines() if ln.strip()]
    urls = [u for u in lines if u.startswith(("http://", "https://")) or "." in u]

    if not urls:
        return jsonify({"error": "No valid URLs"}), 400

    task = CheckTask(urls, delay_ms, lang)
    with tasks_lock:
        tasks[task.id] = task

    thread = threading.Thread(target=run_check, args=(task,), daemon=True)
    thread.start()

    return jsonify({"task_id": task.id, "total": len(urls)})


@app.route("/manual-result/<task_id>", methods=["POST"])
def manual_result(task_id):
    with tasks_lock:
        task = tasks.get(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404

    data = request.get_json()
    status = data.get("status", "error")
    comment = data.get("comment", "")
    title = data.get("title", "")
    idx = data.get("index")

    if task.status == "running" and idx is None:
        task.manual_result = {
            "status": status, "title": title, "comment": comment,
        }
        task.waiting_for_manual.set()
        return jsonify({"ok": True})

    if idx is not None and 0 <= idx < len(task.results):
        task.results[idx]["status"] = status
        task.results[idx]["comment"] = comment
        if title:
            task.results[idx]["title"] = title

    return jsonify({"ok": True})


@app.route("/save-results", methods=["POST"])
def save_results():
    data = request.get_json()
    results = data.get("results", [])
    if not results:
        return jsonify({"error": "No results"}), 400

    task = CheckTask([], 0, "uk")
    task.results = results
    task.status = "done"
    with tasks_lock:
        tasks[task.id] = task

    return jsonify({"task_id": task.id})


@app.route("/stop/<task_id>", methods=["POST"])
def stop_check(task_id):
    with tasks_lock:
        task = tasks.get(task_id)
    if task:
        task.stop_event.set()
        task.waiting_for_manual.set()
        return jsonify({"ok": True})
    return jsonify({"error": "Task not found"}), 404


@app.route("/events/<task_id>")
def events(task_id):
    with tasks_lock:
        task = tasks.get(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404

    def stream():
        idx = 0
        while True:
            while idx < len(task.events):
                evt = task.events[idx]
                yield f"data: {json.dumps(evt, ensure_ascii=False)}\n\n"
                idx += 1
                if evt.get("type") == "done":
                    return
            time.sleep(0.3)

    return Response(
        stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.route("/export/<task_id>/<fmt>")
def export(task_id, fmt):
    with tasks_lock:
        task = tasks.get(task_id)
    if not task or not task.results:
        return jsonify({"error": "No results"}), 404

    label_map = {
        "indexed": "В індексі",
        "not_indexed": "Не в індексі",
        "blocked": "Заблоковано",
        "error": "Помилка",
        "skip": "Пропущено",
    }

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["№", "URL", "Статус", "Заголовок", "Коментар"])
        for r in task.results:
            writer.writerow([
                r.get("num", ""), r.get("url", ""),
                label_map.get(r.get("status", ""), r.get("status", "")),
                r.get("title", ""), r.get("comment", ""),
            ])
        output = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
        output.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(output, mimetype="text/csv", as_attachment=True,
                         download_name=f"index_check_{ts}.csv")

    elif fmt == "xlsx" and OPENPYXL_OK:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результати"
        headers = ["№", "URL", "Статус", "Заголовок", "Коментар"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1A237E", end_color="1A237E",
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        fill_map = {
            "indexed": "C8E6C9", "not_indexed": "FFCDD2",
            "blocked": "FFE0B2", "error": "FFF9C4", "skip": "E0E0E0",
        }
        for r in task.results:
            ws.append([
                r.get("num", ""), r.get("url", ""),
                label_map.get(r.get("status", ""), r.get("status", "")),
                r.get("title", ""), r.get("comment", ""),
            ])
            color = fill_map.get(r.get("status", ""), "F5F5F5")
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill(start_color=color, end_color=color,
                                        fill_type="solid")

        col_widths = [6, 70, 15, 50, 50]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"index_check_{ts}.xlsx",
        )

    return jsonify({"error": "Unsupported format"}), 400


@app.route("/health")
def health():
    """Diagnostic endpoint — is Chromium alive?"""
    info = {
        "pw_ready": _pw_ready.is_set(),
        "pw_alive": _pw_alive,
        "pw_thread": (_pw_thread is not None and _pw_thread.is_alive())
                     if _pw_thread else False,
    }
    if _pw_alive and _pw_thread and _pw_thread.is_alive():
        holder = {}
        done = threading.Event()
        _pw_queue.put(("google.com", "uk", holder, done))
        done.wait(timeout=90)
        result = holder.get("result")
        if result:
            info["test_status"] = result[0]
            info["test_comment"] = result[2]
        else:
            info["test_status"] = "timeout"
    else:
        info["test_status"] = "worker_not_running"
        _ensure_pw_worker()
    return jsonify(info)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port, threaded=True)
